#!/usr/bin/env python3
"""
Test Simulator – Question Pipeline CLI

Usage:
  python main.py load --file templates/unprocessed/questions.xlsx
  python main.py load --dir  templates/unprocessed/
  python main.py load --file templates/unprocessed/questions.xlsx --dry-run
  python main.py validate --file templates/unprocessed/questions.xlsx
  python main.py create-template --out templates/unprocessed/my_questions.xlsx
"""

import os
import sys
from pathlib import Path

import click
from dotenv import load_dotenv
from rich.console import Console
from rich.table import Table

load_dotenv(Path(__file__).parent.parent / ".env")

# Allow imports from pipeline/processors/
sys.path.insert(0, str(Path(__file__).parent))

from processors.excel_parser import parse_excel
from processors.question_validator import validate_rows
from processors.db_loader import load_rows

console = Console()


def _get_db_url() -> str:
    url = os.getenv("DATABASE_URL", "")
    if not url:
        console.print("[red]DATABASE_URL not set. Add it to .env or export it.[/red]")
        sys.exit(1)
    return url


def _process_file(file_path: Path, dry_run: bool, db_url: str) -> dict:
    console.print(f"\n[bold cyan]Processing:[/bold cyan] {file_path.name}")

    rows, parse_errors = parse_excel(file_path)
    for e in parse_errors:
        console.print(f"  [red]Parse error:[/red] {e}")

    if not rows:
        console.print("  [yellow]No rows to process.[/yellow]")
        return {"parsed": 0, "valid": 0, "inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

    valid_rows, warnings = validate_rows(rows)
    for w in warnings:
        console.print(f"  [yellow]Warning:[/yellow] {w}")

    console.print(f"  Parsed: {len(rows)}, Valid: {len(valid_rows)}")

    if valid_rows and not dry_run:
        stats = load_rows(database_url=db_url, rows=valid_rows, dry_run=False)
    elif valid_rows and dry_run:
        stats = load_rows(database_url=db_url, rows=valid_rows, dry_run=True)
        console.print("  [dim](dry-run – no changes committed)[/dim]")
    else:
        stats = {"inserted": 0, "updated": 0, "skipped": len(rows) - len(valid_rows), "errors": 0}

    return {"parsed": len(rows), "valid": len(valid_rows), **stats}


@click.group()
def cli():
    """Test Simulator question pipeline."""


@cli.command()
@click.option("--file", "-f", type=click.Path(exists=True, path_type=Path), help="Single Excel file to load")
@click.option("--dir",  "-d", type=click.Path(exists=True, file_okay=False, path_type=Path), help="Directory of Excel files")
@click.option("--dry-run", is_flag=True, default=False, help="Parse and validate without writing to DB")
def load(file: Path | None, dir: Path | None, dry_run: bool):
    """Load questions from Excel file(s) into the database."""
    db_url = _get_db_url()
    files: list[Path] = []

    if file:
        files.append(file)
    elif dir:
        files = sorted(dir.glob("*.xlsx")) + sorted(dir.glob("*.xls"))
    else:
        console.print("[red]Provide --file or --dir[/red]")
        sys.exit(1)

    if not files:
        console.print("[yellow]No Excel files found.[/yellow]")
        return

    totals = {"parsed": 0, "valid": 0, "inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
    for f in files:
        stats = _process_file(f, dry_run=dry_run, db_url=db_url)
        for k in totals:
            totals[k] += stats.get(k, 0)

    table = Table(title="Pipeline Summary", show_header=True, header_style="bold magenta")
    table.add_column("Metric", style="cyan")
    table.add_column("Count", justify="right")
    for k, v in totals.items():
        color = "green" if k == "inserted" else "yellow" if k == "updated" else "red" if k == "errors" else "white"
        table.add_row(k.capitalize(), f"[{color}]{v}[/{color}]")

    console.print()
    console.print(table)
    if dry_run:
        console.print("\n[bold yellow]DRY RUN – no changes were committed.[/bold yellow]")
    else:
        console.print("\n[bold green]Done![/bold green]")


@cli.command()
@click.option("--file", "-f", required=True, type=click.Path(exists=True, path_type=Path))
def validate(file: Path):
    """Validate an Excel file without loading it."""
    rows, errors = parse_excel(file)
    for e in errors:
        console.print(f"[red]Error:[/red] {e}")
    valid_rows, warnings = validate_rows(rows)
    for w in warnings:
        console.print(f"[yellow]Warning:[/yellow] {w}")
    console.print(f"\n[green]Valid rows: {len(valid_rows)} / {len(rows)}[/green]")


@cli.command("create-template")
@click.option("--out", "-o", default="templates/unprocessed/questions_template.xlsx",
              type=click.Path(path_type=Path), show_default=True)
def create_template(out: Path):
    """Generate a blank Excel template for entering questions."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        console.print("[red]openpyxl not installed. Run: pip install openpyxl[/red]")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "test_type", "subject", "grade", "topic", "question", "passage",
        "type", "difficulty", "points", "explanation",
        "option_a", "option_b", "option_c", "option_d", "option_e", "correct",
    ]
    notes = [
        "nwea_map | math_olympiad | science_olympiad",
        "math | reading | language | science | comp_math | …",
        "K | 1 | 2 | … | 12",
        "Topic name (e.g. Fractions)",
        "Question text (can be long)",
        "Reading passage (optional)",
        "single | multiple",
        "easy | medium | hard",
        "1",
        "Explanation shown after test (optional)",
        "Option A text", "Option B text", "Option C text (optional)",
        "Option D text (optional)", "Option E text (optional)",
        'Correct labels: "A" or "A,C"',
    ]
    sample = [
        "nwea_map", "math", "6", "Ratios & Proportional Relationships",
        "A recipe uses 3 cups of flour for every 2 cups of sugar. If you use 9 cups of flour, how many cups of sugar do you need?",
        "",
        "single", "easy", "1",
        "Set up the proportion: 3/2 = 9/x → x = 6",
        "4 cups", "6 cups", "8 cups", "12 cups", "", "B",
    ]

    header_fill = PatternFill("solid", fgColor="0369A1")
    note_fill   = PatternFill("solid", fgColor="E0F2FE")
    sample_fill = PatternFill("solid", fgColor="F0FDF4")

    for col_idx, (header, note, sample_val) in enumerate(zip(headers, notes, sample), start=1):
        h_cell = ws.cell(row=1, column=col_idx, value=header)
        h_cell.font = Font(bold=True, color="FFFFFF")
        h_cell.fill = header_fill
        h_cell.alignment = Alignment(horizontal="center")

        n_cell = ws.cell(row=2, column=col_idx, value=note)
        n_cell.fill = note_fill
        n_cell.font = Font(italic=True, size=8, color="0369A1")

        s_cell = ws.cell(row=3, column=col_idx, value=sample_val)
        s_cell.fill = sample_fill

        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(15, len(header) + 4)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    console.print(f"[green]Template created:[/green] {out}")
    console.print("\nFill rows starting from row 4. Row 2 shows allowed values; row 3 is a sample.")


if __name__ == "__main__":
    cli()
